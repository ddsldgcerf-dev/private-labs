from fastapi import FastAPI, Request, Form, HTTPException
from fastapi.responses import HTMLResponse, FileResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from starlette.responses import Response

import sqlite3
import os
import json
import hmac
import hashlib
import secrets
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# =========================================================
# BASIC CONFIGURATION
# =========================================================

BASE = Path(__file__).resolve().parent
DB = BASE / "labs.db"

templates = Jinja2Templates(directory=str(BASE / "templates"))

app = FastAPI(title="قاعدة بيانات المخابر الخاصة")


# =========================================================
# REGIONS / OFFICIAL WILAYA NUMBERS
# =========================================================

REGIONS = {
    "الجزائر": ["الجزائر", "بومرداس", "تيبازة"],
    "البليدة": ["البليدة", "البويرة", "تيزي وزو", "المدية", "عين الدفلى", "الجلفة"],
    "عنابة": ["عنابة", "سكيكدة", "سوق أهراس", "الطارف", "قالمة"],
    "وهران": ["وهران", "تلمسان", "سيدي بلعباس", "عين تموشنت", "مستغانم"],
    "بشار": ["بشار", "النعامة", "تندوف", "أدرار", "البيض", "بني عباس", "تيميمون", "برج باجي مختار"],
    "سطيف": ["سطيف", "جيجل", "برج بوعريريج", "المسيلة", "ميلة", "بجاية"],
    "باتنة": ["باتنة", "قسنطينة", "أم البواقي", "تبسة", "خنشلة", "بسكرة", "أولاد جلال"],
    "سعيدة": ["سعيدة", "تيارت", "غليزان", "الشلف", "تيسمسيلت", "معسكر"],
    "ورقلة": ["ورقلة", "غرداية", "الوادي", "إليزي", "تمنراست", "الأغواط",
             "تقرت", "المغير", "المنيعة", "جانت", "عين صالح", "عين قزام"],
}

# الأرقام الرسمية للولايات 01 إلى 58
WILAYA_NUMBERS = {
    "أدرار": 1, "الشلف": 2, "الأغواط": 3, "أم البواقي": 4, "باتنة": 5,
    "بجاية": 6, "بسكرة": 7, "بشار": 8, "البليدة": 9, "البويرة": 10,
    "تمنراست": 11, "تبسة": 12, "تلمسان": 13, "تيارت": 14, "تيزي وزو": 15,
    "الجزائر": 16, "الجلفة": 17, "جيجل": 18, "سطيف": 19, "سعيدة": 20,
    "سكيكدة": 21, "سيدي بلعباس": 22, "عنابة": 23, "قالمة": 24,
    "قسنطينة": 25, "المدية": 26, "مستغانم": 27, "المسيلة": 28,
    "معسكر": 29, "ورقلة": 30, "وهران": 31, "البيض": 32, "إليزي": 33,
    "برج بوعريريج": 34, "بومرداس": 35, "الطارف": 36, "تندوف": 37,
    "تيسمسيلت": 38, "الوادي": 39, "خنشلة": 40, "سوق أهراس": 41,
    "تيبازة": 42, "ميلة": 43, "عين الدفلى": 44, "النعامة": 45,
    "عين تموشنت": 46, "غرداية": 47, "غليزان": 48, "تيميمون": 49, "برج باجي مختار": 50,
    "أولاد جلال": 51, "بني عباس": 52, "عين صالح": 53, "عين قزام": 54,
    "تقرت": 55, "جانت": 56, "المغير": 57, "المنيعة": 58,
}

# هذا القاموس هو المرجع الوحيد لتسمية حسابات DCW، وفق الترقيم الرسمي للولايات 01 إلى 58.
ALL_WILAYAS = {
    wilaya: region
    for region, wilayas in REGIONS.items()
    for wilaya in wilayas
}

HEADERS = [
    "المديرية الجهوية", "المديرية الولائية", "اسم المخبر",
    "رقم السجل التجاري", "رقم رخصة الاستغلال", "تاريخ توقيع الرخصة",
    "عنوان النشاط", "صاحب الرخصة", "المدير التقني", "وضعية المخبر",
    "اختصاص النشاط", "عائلات المنتوجات",
    "التحاليل الفيزيائية والكيميائية", "التحاليل الميكروبيولوجية",
    "تاريخ آخر تفتيش", "ملاحظات",
]


# =========================================================
# DATABASE
# =========================================================

def db():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    return con


def column_exists(con, table, column):
    rows = con.execute(f"PRAGMA table_info({table})").fetchall()
    return any(row["name"] == column for row in rows)


def init_db():
    con = db()

    con.execute("""
        CREATE TABLE IF NOT EXISTS labs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            region TEXT NOT NULL,
            wilaya TEXT NOT NULL,
            name TEXT NOT NULL,
            rc TEXT,
            license_no TEXT,
            license_date TEXT,
            address TEXT,
            owner TEXT,
            technical_manager TEXT,
            status TEXT,
            activity_specialty TEXT,
            product_families TEXT,
            physico_chemical TEXT,
            microbiological TEXT,
            last_inspection TEXT,
            notes TEXT
        )
    """)

    for column in ("activity_specialty", "product_families"):
        if not column_exists(con, "labs", column):
            con.execute(f"ALTER TABLE labs ADD COLUMN {column} TEXT")

    con.execute("""
        CREATE TABLE IF NOT EXISTS accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            region TEXT,
            wilaya TEXT,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    con.commit()
    con.close()


init_db()


# =========================================================
# PASSWORDS
# =========================================================

def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    key = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt, 200_000
    )
    return "pbkdf2_sha256$" + salt.hex() + "$" + key.hex()


def verify_password(password: str, stored: str) -> bool:
    try:
        algorithm, salt_hex, key_hex = stored.split("$")
        if algorithm != "pbkdf2_sha256":
            return False
        salt = bytes.fromhex(salt_hex)
        key = hashlib.pbkdf2_hmac(
            "sha256", password.encode("utf-8"), salt, 200_000
        )
        return hmac.compare_digest(key.hex(), key_hex)
    except Exception:
        return False


def generate_initial_password():
    return secrets.token_urlsafe(9)


# =========================================================
# SESSION
# =========================================================

SESSION_SECRET = os.getenv("SESSION_SECRET", "CHANGE_THIS_SESSION_SECRET")


def sign_session(data: dict) -> str:
    payload = json.dumps(
        data, ensure_ascii=False, separators=(",", ":")
    ).encode()
    encoded = payload.hex()
    signature = hmac.new(
        SESSION_SECRET.encode(), encoded.encode(), hashlib.sha256
    ).hexdigest()
    return encoded + "." + signature


def read_session(request: Request):
    token = request.cookies.get("session")
    if not token or "." not in token:
        return None

    encoded, signature = token.rsplit(".", 1)
    expected = hmac.new(
        SESSION_SECRET.encode(), encoded.encode(), hashlib.sha256
    ).hexdigest()

    if not hmac.compare_digest(signature, expected):
        return None

    try:
        return json.loads(bytes.fromhex(encoded).decode())
    except Exception:
        return None


def set_session(response: Response, account: dict):
    token = sign_session(account)
    response.set_cookie(
        key="session",
        value=token,
        httponly=True,
        secure=True,
        samesite="lax",
        max_age=60 * 60 * 12,
    )


def clear_session(response: Response):
    response.delete_cookie("session")


def current_user(request: Request):
    return read_session(request)


def require_login(request: Request):
    return current_user(request)


# =========================================================
# AUTHORIZATION
# =========================================================

def region_allows(user, region):
    if user["role"] == "admin":
        return True
    if user["role"] == "region":
        return region == user["region"]
    if user["role"] == "wilaya":
        return region == ALL_WILAYAS.get(user["wilaya"])
    return False


def wilaya_allows(user, wilaya):
    if user["role"] == "admin":
        return True
    if user["role"] == "region":
        return wilaya in REGIONS.get(user["region"], [])
    if user["role"] == "wilaya":
        return wilaya == user["wilaya"]
    return False


# =========================================================
# FILTERED LABS
# =========================================================

def get_filtered_labs(
    user, region=None, wilaya=None, activity=None,
    products=None, search=None
):
    con = db()

    query = "SELECT * FROM labs WHERE 1=1"
    params = []

    if user["role"] == "region":
        query += " AND region = ?"
        params.append(user["region"])
    elif user["role"] == "wilaya":
        query += " AND wilaya = ?"
        params.append(user["wilaya"])

    if region and user["role"] == "admin":
        query += " AND region = ?"
        params.append(region)

    if wilaya:
        if wilaya_allows(user, wilaya):
            query += " AND wilaya = ?"
            params.append(wilaya)
        else:
            query += " AND 1 = 0"

    if activity:
        query += """
            AND LOWER(COALESCE(activity_specialty, ''))
            LIKE LOWER(?)
        """
        params.append("%" + activity + "%")

    if products:
        query += """
            AND LOWER(COALESCE(product_families, ''))
            LIKE LOWER(?)
        """
        params.append("%" + products + "%")

    if search:
        query += """
            AND (
                LOWER(name) LIKE LOWER(?)
                OR LOWER(owner) LIKE LOWER(?)
                OR LOWER(rc) LIKE LOWER(?)
                OR LOWER(address) LIKE LOWER(?)
            )
        """
        value = "%" + search + "%"
        params.extend([value, value, value, value])

    query += " ORDER BY region, wilaya, name"

    rows = con.execute(query, params).fetchall()

    activity_values = [
        row[0] for row in con.execute("""
            SELECT DISTINCT activity_specialty
            FROM labs
            WHERE activity_specialty IS NOT NULL
              AND TRIM(activity_specialty) != ''
            ORDER BY activity_specialty
        """).fetchall()
    ]

    product_values = [
        row[0] for row in con.execute("""
            SELECT DISTINCT product_families
            FROM labs
            WHERE product_families IS NOT NULL
              AND TRIM(product_families) != ''
            ORDER BY product_families
        """).fetchall()
    ]

    con.close()
    return rows, activity_values, product_values


def filter_context(user, request):
    region_filter = request.query_params.get("region")
    wilaya_filter = request.query_params.get("wilaya")
    activity = request.query_params.get("activity")
    products = request.query_params.get("products")
    search = request.query_params.get("search")

    rows, activity_values, product_values = get_filtered_labs(
        user, region_filter, wilaya_filter,
        activity, products, search
    )

    if user["role"] == "admin":
        available_regions = list(REGIONS.keys())
        available_wilayas = [
            w for ws in REGIONS.values() for w in ws
        ]
    elif user["role"] == "region":
        available_regions = [user["region"]]
        available_wilayas = REGIONS.get(user["region"], [])
    else:
        available_regions = [ALL_WILAYAS.get(user["wilaya"])]
        available_wilayas = [user["wilaya"]]

    return {
        "rows": rows,
        "count": len(rows),
        "activity_values": activity_values,
        "product_values": product_values,
        "available_regions": available_regions,
        "available_wilayas": available_wilayas,
        "selected_region": region_filter or "",
        "selected_wilaya": wilaya_filter or "",
        "selected_activity": activity or "",
        "selected_products": products or "",
        "selected_search": search or "",
        "user": user,
    }


# =========================================================
# LOGIN
# =========================================================

@app.get("/", response_class=HTMLResponse)
def login_page(request: Request):
    user = current_user(request)

    if user:
        if user["role"] == "admin":
            return RedirectResponse("/admin", status_code=303)
        if user["role"] == "region":
            return RedirectResponse("/region", status_code=303)
        if user["role"] == "wilaya":
            return RedirectResponse("/wilaya", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="login.html",
        context={"request": request, "error": None},
    )


@app.post("/login")
def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...)
):
    username = username.strip()

    admin_password = os.getenv("ADMIN_PASSWORD", "")

    if username == "admin" and admin_password:
        if hmac.compare_digest(password, admin_password):
            response = RedirectResponse("/admin", status_code=303)
            set_session(response, {
                "username": "admin",
                "role": "admin",
                "region": None,
                "wilaya": None,
            })
            return response

    con = db()
    account = con.execute("""
        SELECT * FROM accounts
        WHERE username = ? AND active = 1
    """, (username,)).fetchone()
    con.close()

    if account and verify_password(password, account["password_hash"]):
        destination = "/region" if account["role"] == "region" else "/wilaya"
        response = RedirectResponse(destination, status_code=303)
        set_session(response, {
            "username": account["username"],
            "role": account["role"],
            "region": account["region"],
            "wilaya": account["wilaya"],
            "account_id": account["id"],
        })
        return response

    return templates.TemplateResponse(
        request=request,
        name="login.html",
        context={
            "request": request,
            "error": "اسم المستخدم أو كلمة المرور غير صحيحة",
        },
        status_code=401,
    )


@app.get("/logout")
def logout():
    response = RedirectResponse("/", status_code=303)
    clear_session(response)
    return response


# =========================================================
# ADMIN
# =========================================================

@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request):
    user = require_login(request)

    if not user:
        return RedirectResponse("/", status_code=303)

    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")

    context = filter_context(user, request)

    return templates.TemplateResponse(
        request=request,
        name="admin.html",
        context={"request": request, **context},
    )


# =========================================================
# CREATE BASIC ACCOUNTS
# =========================================================

@app.post("/admin/accounts/bootstrap")
def bootstrap_accounts(request: Request):
    user = require_login(request)

    if not user or user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")

    con = db()
    created_accounts = []
    existing_accounts = []

    # -----------------------------------------------------
    # DRC ACCOUNTS
    # النمط: 01_drc, 02_drc, ...
    # -----------------------------------------------------

    for index, region in enumerate(REGIONS.keys(), start=1):
        username = f"{index:02d}_drc"

        existing = con.execute("""
            SELECT id, username
            FROM accounts
            WHERE username = ?
        """, (username,)).fetchone()

        if existing:
            existing_accounts.append({
                "username": username,
                "role": "region",
                "region": region,
                "wilaya": None,
            })
            continue

        password = generate_initial_password()

        con.execute("""
            INSERT INTO accounts
            (username, password_hash, role, region, wilaya)
            VALUES (?, ?, ?, ?, ?)
        """, (
            username,
            hash_password(password),
            "region",
            region,
            None,
        ))

        created_accounts.append({
            "username": username,
            "password": password,
            "role": "region",
            "region": region,
            "wilaya": None,
        })

    # -----------------------------------------------------
    # DCW ACCOUNTS
    # النمط: dcw_16 للجزائر، dcw_17 للجلفة... إلخ
    # الرقم مأخوذ من WILAYA_NUMBERS وليس من ترتيب القائمة.
    # -----------------------------------------------------

    for wilaya, region in ALL_WILAYAS.items():
        number = WILAYA_NUMBERS.get(wilaya)

        if number is None:
            # حماية إضافية: لا ننشئ حساباً باسم خاطئ
            continue

        username = f"dcw_{number:02d}"

        existing = con.execute("""
            SELECT id, username
            FROM accounts
            WHERE username = ?
        """, (username,)).fetchone()

        if existing:
            existing_accounts.append({
                "username": username,
                "role": "wilaya",
                "region": region,
                "wilaya": wilaya,
            })
            continue

        password = generate_initial_password()

        con.execute("""
            INSERT INTO accounts
            (username, password_hash, role, region, wilaya)
            VALUES (?, ?, ?, ?, ?)
        """, (
            username,
            hash_password(password),
            "wilaya",
            region,
            wilaya,
        ))

        created_accounts.append({
            "username": username,
            "password": password,
            "role": "wilaya",
            "region": region,
            "wilaya": wilaya,
        })

    con.commit()
    con.close()

    return {
        "success": True,
        "created": created_accounts,
        "existing": existing_accounts,
        "created_count": len(created_accounts),
        "existing_count": len(existing_accounts),
        "message": "تم إنشاء الحسابات الأساسية بنجاح",
    }


# =========================================================
# ADD ACCOUNT MANUALLY
# =========================================================

@app.post("/admin/accounts/add")
def add_account(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form(...),
    region: str = Form(""),
    wilaya: str = Form("")
):
    user = require_login(request)

    if not user or user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")

    username = username.strip()

    if role not in ["region", "wilaya"]:
        raise HTTPException(status_code=400, detail="نوع الحساب غير صحيح")

    if role == "region":
        if region not in REGIONS:
            raise HTTPException(
                status_code=400,
                detail="المديرية الجهوية غير صحيحة"
            )
        wilaya = None

    if role == "wilaya":
        if wilaya not in ALL_WILAYAS:
            raise HTTPException(
                status_code=400,
                detail="المديرية الولائية غير صحيحة"
            )
        region = ALL_WILAYAS[wilaya]

    if len(password) < 6:
        raise HTTPException(
            status_code=400,
            detail="كلمة المرور يجب أن تكون 6 أحرف على الأقل"
        )

    con = db()

    try:
        con.execute("""
            INSERT INTO accounts
            (username, password_hash, role, region, wilaya)
            VALUES (?, ?, ?, ?, ?)
        """, (
            username,
            hash_password(password),
            role,
            region,
            wilaya,
        ))
        con.commit()
    except sqlite3.IntegrityError:
        con.close()
        raise HTTPException(
            status_code=400,
            detail="اسم المستخدم موجود مسبقًا"
        )

    con.close()
    return RedirectResponse("/admin", status_code=303)


# =========================================================
# CHANGE PASSWORD
# =========================================================

@app.post("/admin/accounts/{account_id}/password")
def change_account_password(
    account_id: int,
    request: Request,
    password: str = Form(...)
):
    user = require_login(request)

    if not user or user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")

    if len(password) < 6:
        raise HTTPException(
            status_code=400,
            detail="كلمة المرور يجب أن تكون 6 أحرف على الأقل"
        )

    con = db()
    con.execute("""
        UPDATE accounts
        SET password_hash = ?
        WHERE id = ?
    """, (hash_password(password), account_id))
    con.commit()
    con.close()

    return RedirectResponse("/admin", status_code=303)


# =========================================================
# ENABLE / DISABLE ACCOUNT
# =========================================================

@app.post("/admin/accounts/{account_id}/toggle")
def toggle_account(account_id: int, request: Request):
    user = require_login(request)

    if not user or user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")

    con = db()
    con.execute("""
        UPDATE accounts
        SET active = CASE WHEN active = 1 THEN 0 ELSE 1 END
        WHERE id = ?
    """, (account_id,))
    con.commit()
    con.close()

    return RedirectResponse("/admin", status_code=303)


# =========================================================
# DELETE ACCOUNT
# =========================================================

@app.post("/admin/accounts/{account_id}/delete")
def delete_account(account_id: int, request: Request):
    user = require_login(request)

    if not user or user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")

    con = db()
    con.execute("DELETE FROM accounts WHERE id = ?", (account_id,))
    con.commit()
    con.close()

    return RedirectResponse("/admin", status_code=303)


# =========================================================
# ACCOUNTS LIST
# =========================================================

@app.get("/admin/accounts")
def accounts_list(request: Request):
    user = require_login(request)

    if not user or user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")

    con = db()
    accounts = con.execute("""
        SELECT id, username, role, region, wilaya, active, created_at
        FROM accounts
        ORDER BY role, region, wilaya, username
    """).fetchall()
    con.close()

    return {"accounts": [dict(row) for row in accounts]}


# =========================================================
# REGIONAL DASHBOARD
# =========================================================

@app.get("/region", response_class=HTMLResponse)
def region_page(request: Request):
    user = require_login(request)

    if not user:
        return RedirectResponse("/", status_code=303)

    if user["role"] != "region":
        raise HTTPException(status_code=403, detail="غير مصرح")

    context = filter_context(user, request)

    return templates.TemplateResponse(
        request=request,
        name="region.html",
        context={"request": request, **context},
    )


# =========================================================
# WILAYA DASHBOARD
# =========================================================

@app.get("/wilaya", response_class=HTMLResponse)
def wilaya_page(request: Request):
    user = require_login(request)

    if not user:
        return RedirectResponse("/", status_code=303)

    if user["role"] != "wilaya":
        raise HTTPException(status_code=403, detail="غير مصرح")

    context = filter_context(user, request)

    return templates.TemplateResponse(
        request=request,
        name="wilaya.html",
        context={"request": request, **context},
    )


# =========================================================
# ADD LAB
# =========================================================

@app.post("/add")
def add_lab(
    request: Request,
    region: str = Form(...),
    wilaya: str = Form(...),
    name: str = Form(...),
    rc: str = Form(""),
    license_no: str = Form(""),
    license_date: str = Form(""),
    address: str = Form(""),
    owner: str = Form(""),
    technical_manager: str = Form(""),
    status: str = Form("نشط"),
    activity_specialty: str = Form(""),
    product_families: str = Form(""),
    physico_chemical: str = Form(""),
    microbiological: str = Form(""),
    last_inspection: str = Form(""),
    notes: str = Form("")
):
    user = require_login(request)

    if not user:
        return RedirectResponse("/", status_code=303)

    if user["role"] != "wilaya":
        raise HTTPException(
            status_code=403,
            detail="الإدارة المركزية والمديرية الجهوية لا تضيف البيانات من هذه الصفحة"
        )

    # لا نثق بالقيم القادمة من النموذج؛ نأخذ نطاق المستخدم من الجلسة
    region = user["region"]
    wilaya = user["wilaya"]

    if region not in REGIONS:
        raise HTTPException(
            status_code=400,
            detail="المديرية الجهوية غير صحيحة"
        )

    if wilaya not in REGIONS[region]:
        raise HTTPException(
            status_code=400,
            detail="المديرية الولائية لا تتبع المديرية الجهوية"
        )

    con = db()

    con.execute("""
        INSERT INTO labs (
            region, wilaya, name, rc, license_no, license_date,
            address, owner, technical_manager, status,
            activity_specialty, product_families,
            physico_chemical, microbiological,
            last_inspection, notes
        )
        VALUES (
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
        )
    """, (
        region, wilaya, name, rc, license_no, license_date,
        address, owner, technical_manager, status,
        activity_specialty, product_families,
        physico_chemical, microbiological,
        last_inspection, notes,
    ))

    con.commit()
    con.close()

    return RedirectResponse("/wilaya?sent=1", status_code=303)


# =========================================================
# DELETE LAB - ADMIN ONLY
# =========================================================

@app.post("/admin/labs/{lab_id}/delete")
def delete_lab(lab_id: int, request: Request):
    user = require_login(request)

    if not user or user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")

    con = db()
    con.execute("DELETE FROM labs WHERE id = ?", (lab_id,))
    con.commit()
    con.close()

    return RedirectResponse("/admin", status_code=303)


# =========================================================
# EXPORT EXCEL
# =========================================================

@app.get("/admin/export")
def export_excel(request: Request):
    user = require_login(request)

    if not user or user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")

    context = filter_context(user, request)
    rows = context["rows"]

    wb = Workbook()
    ws = wb.active
    ws.title = "قاعدة المخابر"
    ws.sheet_view.rightToLeft = True

    ws.append(HEADERS)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    fields = [
        "region", "wilaya", "name", "rc", "license_no",
        "license_date", "address", "owner", "technical_manager",
        "status", "activity_specialty", "product_families",
        "physico_chemical", "microbiological",
        "last_inspection", "notes",
    ]

    for row in rows:
        ws.append([row[field] for field in fields])

    for col in ws.columns:
        max_len = max(
            len(str(cell.value or ""))
            for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(
            max(max_len + 3, 12), 45
        )

    path = BASE / "قاعدة_المخابر_الخاصة.xlsx"
    wb.save(path)

    return FileResponse(
        path,
        filename="قاعدة_المخابر_الخاصة.xlsx",
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


# =========================================================
# INSPECTIONS / INSPECTION RETURNS
# إضافة مستقلة: لا تمس قاعدة بيانات المخابر أو صلاحياتها
# =========================================================

INSPECTION_HEADERS = [
    "المديرية الجهوية",
    "المديرية الولائية",
    "اسم المخبر",
    "تاريخ التفتيش",
    "هل توجد مخالفات؟",
    "المخالفات المرصودة",
    "هل توجد مخالفات سابقة؟",
    "وضعية المخالفات السابقة",
    "هل تم رفع التحفظ؟",
    "هل تمت معالجة المخالفات؟",
    "ملاحظات",
]


def init_inspections_db():
    """
    قاعدة مستقلة لحصائل عمليات التفتيش.

    لا تتضمن:
        - اسم المفتش
        - أعضاء فريق التفتيش
        - نوع التفتيش
        - سبب التفتيش
        - عينات
        - منتجات
        - سحب أو حجز أو إتلاف

    كل عملية تفتيش تسجل كعملية مستقلة، ومن ثم يمكن استخراج
    الحصائل الشهرية والثلاثية ونصف السنوية والسنوية بالاستعلامات.
    """
    con = db()

    con.execute("""
        CREATE TABLE IF NOT EXISTS inspections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            region TEXT NOT NULL,
            wilaya TEXT NOT NULL,
            lab_name TEXT NOT NULL,
            inspection_date TEXT NOT NULL,
            has_violations INTEGER NOT NULL DEFAULT 0,
            violations TEXT,
            previous_violations INTEGER NOT NULL DEFAULT 0,
            previous_violations_status TEXT,
            reservations_lifted INTEGER NOT NULL DEFAULT 0,
            violations_processed INTEGER NOT NULL DEFAULT 0,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    con.commit()
    con.close()


init_inspections_db()


def inspection_user_allows(user, region=None, wilaya=None):
    """نفس نطاق الصلاحيات المعتمد في قاعدة المخابر."""
    if not user:
        return False

    if user["role"] == "admin":
        return True

    if user["role"] == "region":
        if region is not None and region != user["region"]:
            return False
        if wilaya is not None and not wilaya_allows(user, wilaya):
            return False
        return True

    if user["role"] == "wilaya":
        if wilaya is not None:
            return wilaya == user["wilaya"]
        if region is not None:
            return region == user["region"]
        return True

    return False


def get_filtered_inspections(
    user,
    region=None,
    wilaya=None,
    date_from=None,
    date_to=None,
    has_violations=None,
    previous_violations=None,
    reservations_lifted=None,
    violations_processed=None,
    search=None,
):
    """استخراج سجل التفتيش وفق نطاق المستخدم وجميع المرشحات المتاحة."""

    con = db()

    query = "SELECT * FROM inspections WHERE 1=1"
    params = []

    # -----------------------------------------------------
    # SECURITY SCOPE
    # -----------------------------------------------------

    if user["role"] == "region":
        query += " AND region = ?"
        params.append(user["region"])

    elif user["role"] == "wilaya":
        query += " AND wilaya = ?"
        params.append(user["wilaya"])

    # -----------------------------------------------------
    # LOCATION FILTERS
    # -----------------------------------------------------

    if region:
        if region_allows(user, region):
            query += " AND region = ?"
            params.append(region)
        else:
            query += " AND 1 = 0"

    if wilaya:
        if wilaya_allows(user, wilaya):
            query += " AND wilaya = ?"
            params.append(wilaya)
        else:
            query += " AND 1 = 0"

    # -----------------------------------------------------
    # DATE FILTERS
    # -----------------------------------------------------

    if date_from:
        query += " AND inspection_date >= ?"
        params.append(date_from)

    if date_to:
        query += " AND inspection_date <= ?"
        params.append(date_to)

    # -----------------------------------------------------
    # VIOLATIONS FILTERS
    # -----------------------------------------------------

    if has_violations in ("0", "1"):
        query += " AND has_violations = ?"
        params.append(int(has_violations))

    if previous_violations in ("0", "1"):
        query += " AND previous_violations = ?"
        params.append(int(previous_violations))

    if reservations_lifted in ("0", "1"):
        query += " AND reservations_lifted = ?"
        params.append(int(reservations_lifted))

    if violations_processed in ("0", "1"):
        query += " AND violations_processed = ?"
        params.append(int(violations_processed))

    # -----------------------------------------------------
    # LAB SEARCH
    # -----------------------------------------------------

    if search:
        query += """
            AND (
                LOWER(lab_name) LIKE LOWER(?)
                OR LOWER(violations) LIKE LOWER(?)
                OR LOWER(previous_violations_status) LIKE LOWER(?)
                OR LOWER(notes) LIKE LOWER(?)
            )
        """
        value = "%" + search + "%"
        params.extend([value, value, value, value])

    query += " ORDER BY inspection_date DESC, region, wilaya, lab_name"

    rows = con.execute(query, params).fetchall()
    con.close()

    return rows


# =========================================================
# ADD INSPECTION RETURN - DCW ONLY
# =========================================================

@app.post("/inspections/add")
def add_inspection(
    request: Request,
    lab_name: str = Form(...),
    inspection_date: str = Form(...),
    has_violations: int = Form(0),
    violations: str = Form(""),
    previous_violations: int = Form(0),
    previous_violations_status: str = Form(""),
    reservations_lifted: int = Form(0),
    violations_processed: int = Form(0),
    notes: str = Form(""),
):
    user = require_login(request)

    if not user:
        return RedirectResponse("/", status_code=303)

    # التسجيل يتم من طرف المديرية الولائية فقط.
    if user["role"] != "wilaya":
        raise HTTPException(
            status_code=403,
            detail="المديرية الولائية فقط لها صلاحية إدخال حصائل عمليات التفتيش"
        )

    region = user["region"]
    wilaya = user["wilaya"]

    if region not in REGIONS:
        raise HTTPException(
            status_code=400,
            detail="المديرية الجهوية غير صحيحة"
        )

    if wilaya not in REGIONS[region]:
        raise HTTPException(
            status_code=400,
            detail="المديرية الولائية لا تتبع المديرية الجهوية"
        )

    if not lab_name.strip():
        raise HTTPException(
            status_code=400,
            detail="اسم المخبر مطلوب"
        )

    if has_violations not in (0, 1):
        raise HTTPException(status_code=400, detail="قيمة المخالفات غير صحيحة")

    if previous_violations not in (0, 1):
        raise HTTPException(status_code=400, detail="قيمة المخالفات السابقة غير صحيحة")

    if reservations_lifted not in (0, 1):
        raise HTTPException(status_code=400, detail="قيمة رفع التحفظ غير صحيحة")

    if violations_processed not in (0, 1):
        raise HTTPException(status_code=400, detail="قيمة معالجة المخالفات غير صحيحة")

    con = db()

    con.execute("""
        INSERT INTO inspections (
            region,
            wilaya,
            lab_name,
            inspection_date,
            has_violations,
            violations,
            previous_violations,
            previous_violations_status,
            reservations_lifted,
            violations_processed,
            notes
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        region,
        wilaya,
        lab_name.strip(),
        inspection_date,
        has_violations,
        violations.strip(),
        previous_violations,
        previous_violations_status.strip(),
        reservations_lifted,
        violations_processed,
        notes.strip(),
    ))

    con.commit()
    con.close()

    return RedirectResponse("/wilaya?inspection_sent=1", status_code=303)


# =========================================================
# INSPECTIONS VIEW
# DCW: own wilaya / DRC: own region / ADMIN: all
# =========================================================

@app.get("/inspections", response_class=HTMLResponse)
def inspections_page(request: Request):
    user = require_login(request)

    if not user:
        return RedirectResponse("/", status_code=303)

    context = {
        "request": request,
        "user": user,
        "REGIONS": REGIONS,
    }

    # هذه الصفحة تعتمد لاحقًا على inspections.html،
    # ولا تغيّر أي صفحة من الصفحات السابقة.
    return templates.TemplateResponse(
        request=request,
        name="inspections.html",
        context=context,
    )


# =========================================================
# INSPECTION DATA API / FILTERED RECORDS
# =========================================================

@app.get("/api/inspections")
def inspections_api(request: Request):
    user = require_login(request)

    if not user:
        raise HTTPException(status_code=401, detail="غير مصرح")

    rows = get_filtered_inspections(
        user,
        region=request.query_params.get("region"),
        wilaya=request.query_params.get("wilaya"),
        date_from=request.query_params.get("date_from"),
        date_to=request.query_params.get("date_to"),
        has_violations=request.query_params.get("has_violations"),
        previous_violations=request.query_params.get("previous_violations"),
        reservations_lifted=request.query_params.get("reservations_lifted"),
        violations_processed=request.query_params.get("violations_processed"),
        search=request.query_params.get("search"),
    )

    return {
        "count": len(rows),
        "inspections": [dict(row) for row in rows],
    }


# =========================================================
# INSPECTION SUMMARY
# قاعدة الحصيلة الإضافية: المديرية الجهوية + الولائية + عدد التفتيشات
# لا يظهر اسم المخبر في هذا المخرج التجميعي.
# =========================================================

@app.get("/api/inspections/summary")
def inspections_summary_api(request: Request):
    user = require_login(request)

    if not user:
        raise HTTPException(status_code=401, detail="غير مصرح")

    region = request.query_params.get("region")
    wilaya = request.query_params.get("wilaya")
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")

    if region and not region_allows(user, region):
        raise HTTPException(status_code=403, detail="غير مصرح")

    if wilaya and not wilaya_allows(user, wilaya):
        raise HTTPException(status_code=403, detail="غير مصرح")

    con = db()

    query = """
        SELECT
            region,
            wilaya,
            COUNT(*) AS inspections_count,
            SUM(CASE WHEN has_violations = 1 THEN 1 ELSE 0 END) AS inspections_with_violations,
            SUM(CASE WHEN previous_violations = 1 THEN 1 ELSE 0 END) AS inspections_with_previous_violations,
            SUM(CASE WHEN reservations_lifted = 1 THEN 1 ELSE 0 END) AS reservations_lifted_count,
            SUM(CASE WHEN violations_processed = 1 THEN 1 ELSE 0 END) AS violations_processed_count
        FROM inspections
        WHERE 1=1
    """
    params = []

    if user["role"] == "region":
        query += " AND region = ?"
        params.append(user["region"])
    elif user["role"] == "wilaya":
        query += " AND wilaya = ?"
        params.append(user["wilaya"])

    if region:
        query += " AND region = ?"
        params.append(region)

    if wilaya:
        query += " AND wilaya = ?"
        params.append(wilaya)

    if date_from:
        query += " AND inspection_date >= ?"
        params.append(date_from)

    if date_to:
        query += " AND inspection_date <= ?"
        params.append(date_to)

    query += """
        GROUP BY region, wilaya
        ORDER BY region, wilaya
    """

    rows = con.execute(query, params).fetchall()
    con.close()

    return {
        "count": len(rows),
        "summary": [dict(row) for row in rows],
    }


# =========================================================
# INSPECTION EXCEL EXPORT
# =========================================================

# =========================================================
# INSPECTION EXCEL EXPORT
# متاح لجميع الصلاحيات وفق نطاق كل مستخدم
# Wilaya: ولايته فقط
# Region: مديرياته الولائية فقط
# Admin: جميع البيانات
# =========================================================

@app.get("/inspections/export")
@app.get("/admin/inspections/export")
def export_inspections_excel(request: Request):

    user = require_login(request)

    if not user:
        raise HTTPException(
            status_code=401,
            detail="غير مصرح"
        )

    rows = get_filtered_inspections(
        user,
        region=request.query_params.get("region"),
        wilaya=request.query_params.get("wilaya"),
        date_from=request.query_params.get("date_from"),
        date_to=request.query_params.get("date_to"),
        has_violations=request.query_params.get("has_violations"),
        previous_violations=request.query_params.get("previous_violations"),
        reservations_lifted=request.query_params.get("reservations_lifted"),
        violations_processed=request.query_params.get("violations_processed"),
        search=request.query_params.get("search"),
    )

    wb = Workbook()

    ws = wb.active
    ws.title = "حصائل التفتيش"
    ws.sheet_view.rightToLeft = True

    ws.append(INSPECTION_HEADERS)

    for cell in ws[1]:

        cell.font = Font(bold=True)

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    fields = [
        "region",
        "wilaya",
        "lab_name",
        "inspection_date",
        "has_violations",
        "violations",
        "previous_violations",
        "previous_violations_status",
        "reservations_lifted",
        "violations_processed",
        "notes",
    ]

    for row in rows:

        ws.append([
            row[field]
            for field in fields
        ])

    for col in ws.columns:

        max_len = max(
            len(str(cell.value or ""))
            for cell in col
        )

        ws.column_dimensions[
            col[0].column_letter
        ].width = min(
            max(max_len + 3, 12),
            45
        )

    path = BASE / "حصائل_عمليات_التفتيش.xlsx"

    wb.save(path)

    return FileResponse(
        path,
        filename="حصائل_عمليات_التفتيش.xlsx",
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
